#!/usr/bin/env python3
"""
Uppdaterar VM-tipsets resultat automatiskt.

Flöde:
 1. Hämta spelade matcher från football-data.org (eller en mock-fil).
 2. Skriv in målen i facit-fliken "Resultat & tabell" i master-arket.
 3. Räkna om arket med LibreOffice (headless).
 4. Exportera ställning + alla tips till data.json (som webbsidan läser).

Körs av GitHub Actions enligt schema. Token läses från miljövariabeln
FOOTBALL_DATA_TOKEN (lägg in som "Repository secret" på GitHub).
"""
import os, sys, json, argparse, datetime, unicodedata, urllib.request
from openpyxl import load_workbook
from recalc import recalc

MASTER = 'VM-tips-2026-Swedish_7p_4_4_MASTER.xlsm'
OUT_JSON = 'data.json'
API_URL = 'https://api.football-data.org/v4/competitions/WC/matches'

PLAYERS = ['Lasse Svensson','Christer Lövgren','Conny Lantz','Emma Svensson',
 'Hampus Svensson','Mikael Cordes','Noel Holmström','Oliver Jönsson','Philip Wallberg',
 'Ralle Nilsson Hansen','Rayan Lövgren','Tim Israelsson']

# Svenskt lagnamn (som i arket) -> accepterade API-stavningar
TEAM_ALIASES = {
 'Algeriet':['Algeria'], 'Argentina':['Argentina'], 'Australien':['Australia'],
 'Belgien':['Belgium'], 'Bosnien-Herzigovina':['Bosnia and Herzegovina','Bosnia-Herzegovina','Bosnia & Herzegovina'],
 'Brasilien':['Brazil'], 'Colombia':['Colombia'], 'Curaçao':['Curacao','Curaçao'],
 'Czechia':['Czech Republic','Czechia'], 'Ecuador':['Ecuador'], 'Egypten':['Egypt'],
 'Elfenbenskusten':['Ivory Coast','Cote d Ivoire','Côte d’Ivoire','Cote dIvoire'],
 'England':['England'], 'Frankrike':['France'], 'Ghana':['Ghana'], 'Haiti':['Haiti'],
 'Irak':['Iraq'], 'Iran':['Iran'], 'Japan':['Japan'], 'Jordanien':['Jordan'],
 'Kanada':['Canada'], 'Kap Verde':['Cape Verde','Cabo Verde'],
 'Kongo':['DR Congo','Congo DR','Democratic Republic of the Congo','DR Congo (Kinshasa)'],
 'Kroatien':['Croatia'], 'Marocko':['Morocco'], 'Mexiko':['Mexico'],
 'Nederländerna':['Netherlands','Holland'], 'Norge':['Norway'], 'Nya Zeeland':['New Zealand'],
 'Panama':['Panama'], 'Paraguay':['Paraguay'], 'Portugal':['Portugal'], 'Qatar':['Qatar'],
 'Saudiarabien':['Saudi Arabia'], 'Schweiz':['Switzerland'], 'Senegal':['Senegal'],
 'Skottland':['Scotland'], 'Spanien':['Spain'], 'Sverige':['Sweden'],
 'Sydafrika':['South Africa'], 'Sydkorea':['South Korea','Korea Republic','Republic of Korea','Korea'],
 'Tunisien':['Tunisia'], 'Turkiet':['Turkey','Turkiye','Türkiye'], 'Tyskland':['Germany'],
 'USA':['United States','USA','United States of America'], 'Uruguay':['Uruguay'],
 'Uzbekistan':['Uzbekistan'], 'Österrike':['Austria'],
}

def norm(s):
    if not s: return ''
    s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode()
    return ''.join(c for c in s.lower() if c.isalnum())

# normaliserad alias -> svenskt namn
NORM2SE = {}
for se, al in TEAM_ALIASES.items():
    NORM2SE[norm(se)] = se
    for a in al:
        NORM2SE[norm(a)] = se

def resolve(team):
    """team = API-lagobjekt; returnera svenskt namn eller None."""
    for key in ('name','shortName','tla'):
        v = team.get(key) if isinstance(team, dict) else None
        if v and norm(v) in NORM2SE:
            return NORM2SE[norm(v)]
    return None

def fetch_matches(token, mock=None):
    if mock:
        return json.load(open(mock, encoding='utf-8')).get('matches', [])
    req = urllib.request.Request(API_URL, headers={'X-Auth-Token': token})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r).get('matches', [])

def build_results(matches):
    """frozenset({se_home, se_away}) -> {se_team: goals}"""
    by_pair = {}
    for m in matches:
        if m.get('status') != 'FINISHED':
            continue
        h = resolve(m.get('homeTeam', {})); a = resolve(m.get('awayTeam', {}))
        if not h or not a:
            continue
        ft = (m.get('score') or {}).get('fullTime') or {}
        hg, ag = ft.get('home'), ft.get('away')
        if hg is None or ag is None:
            continue
        by_pair[frozenset((h, a))] = {h: int(hg), a: int(ag)}
    return by_pair

def is_num(v): return isinstance(v, (int, float))

def match_rows(ws):
    """Yield (row, matchno, home, away) for both group and knockout match rows."""
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        d = ws.cell(row=r, column=4).value; f = ws.cell(row=r, column=6).value
        if is_num(b) and isinstance(d, str) and isinstance(f, str) and d.strip() and f.strip():
            yield r, int(b), d.strip(), f.strip()

def apply_results(path, results, manual):
    """Write goals into Resultat & tabell. Returns count written."""
    wb = load_workbook(path, keep_vba=True)
    ws = wb['Resultat & tabell']
    written = 0
    for r, no, home, away in match_rows(ws):
        goals = None
        key = frozenset((home, away))
        if key in results:
            d = results[key]; goals = (d[home], d[away])
            # slutspel kan inte sluta oavgjort i arket (vinnaren måste fram).
            # Hoppa över oavgjorda API-resultat för matchnr > 72 — matas in manuellt.
            if no > 72 and goals[0] == goals[1]:
                goals = None
        if str(no) in manual:        # manuell override vinner alltid
            mh, ma = manual[str(no)]; goals = (mh, ma)
        if goals is not None:
            ws.cell(row=r, column=7).value = goals[0]
            ws.cell(row=r, column=9).value = goals[1]
            written += 1
    wb.save(path)
    return written

def extract(path):
    wb = load_workbook(path, data_only=True)
    res = wb['Resultat & tabell']
    ROUNDS = ('Sextondelsfinal','Åttondelsfinal','Kvartsfinal','Semifinal','Bronsmatch','Final')
    matches = {}
    knockout = []
    for r, no, home, away in match_rows(res):
        date = res.cell(row=r, column=3).value
        hg = res.cell(row=r, column=7).value; ag = res.cell(row=r, column=9).value
        sign = res.cell(row=r, column=10).value
        played = is_num(hg) and is_num(ag)
        datestr = date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else (str(date) if date else None)
        if no > 72:
            rnd = None
            for rr in range(r, 0, -1):
                bv = res.cell(row=rr, column=2).value
                if isinstance(bv, str) and bv.strip() in ROUNDS:
                    rnd = bv.strip(); break
            knockout.append({'no': no, 'round': rnd, 'home': home, 'away': away,
                'date': datestr, 'homeGoals': int(hg) if played else None,
                'awayGoals': int(ag) if played else None, 'played': bool(played)})
            continue
        grp = None
        for rr in range(r, 0, -1):
            bv = res.cell(row=rr, column=2).value
            if isinstance(bv, str) and bv.startswith('Grupp '):
                grp = bv.replace('Grupp ', '').strip(); break
        matches[no] = {'no': no, 'group': grp, 'home': home, 'away': away, 'date': datestr,
            'homeGoals': int(hg) if played else None, 'awayGoals': int(ag) if played else None,
            'sign': str(sign) if sign is not None else None, 'played': bool(played)}
    # --- extra poänglogik som arket inte gör av sig självt ---
    KNOWN_TEAMS = set(TEAM_ALIASES.keys())
    RL_PTS = {'Sextondelsfinal': 1, 'Åttondelsfinal': 2, 'Kvartsfinal': 4,
              'Semifinal': 6, 'Bronsmatch': 8, 'Final': 8}
    # flest gjorda/insläppta mål i gruppspelet — kan delas av flera lag (lika)
    scored, conceded = {}, {}
    group_played = 0
    for m in matches.values():
        if m['played']:
            group_played += 1
            scored[m['home']] = scored.get(m['home'], 0) + m['homeGoals']
            scored[m['away']] = scored.get(m['away'], 0) + m['awayGoals']
            conceded[m['home']] = conceded.get(m['home'], 0) + m['awayGoals']
            conceded[m['away']] = conceded.get(m['away'], 0) + m['homeGoals']
    def leaders(d):
        if not d: return set()
        mx = max(d.values()); return {t.lower() for t, v in d.items() if v == mx}
    group_complete = len(matches) > 0 and group_played == len(matches)
    scored_leaders = leaders(scored) if group_complete else set()
    conceded_leaders = leaders(conceded) if group_complete else set()
    # facit-lag per slutspelsrunda (vilka lag som faktiskt nått omgången)
    ko_round = {m['no']: m['round'] for m in knockout}
    facit_by_round = {}
    for m in knockout:
        s = facit_by_round.setdefault(m['round'], set())
        if m['home'] in KNOWN_TEAMS: s.add(m['home'])
        if m['away'] in KNOWN_TEAMS: s.add(m['away'])

    players = []
    ROUND_LABELS = {'Sextondelsfinaler':'Sextondelsfinal','Åttondelsfinaler':'Åttondelsfinal',
                    'Kvartsfinaler':'Kvartsfinal','Semifinaler':'Semifinal',
                    'Bronsmatch':'Bronsmatch','Final':'Final'}
    for name in PLAYERS:
        if name not in wb.sheetnames: continue
        ws = wb[name]; preds = {}; ko = {}
        for r, no, home, away in match_rows(ws):
            hg = ws.cell(row=r, column=7).value; ag = ws.cell(row=r, column=9).value
            sign = ws.cell(row=r, column=10).value; pts = ws.cell(row=r, column=11).value
            pv = int(pts) if is_num(pts) else 0
            entry = {'home': int(hg) if is_num(hg) else None,
                     'away': int(ag) if is_num(ag) else None,
                     'sign': str(sign) if sign is not None else None, 'points': pv}
            if no > 72:
                entry['predHome'] = home; entry['predAway'] = away
                ko[no] = entry
            else:
                preds[no] = entry
        # per-runda-sammanställning (rad 22-29: X=label, Y=poäng, Z=bonus, AA=totalt)
        rounds = []
        for r in range(22, 30):
            lbl = ws.cell(row=r, column=24).value
            if not (isinstance(lbl, str) and lbl.strip()):
                continue
            y = ws.cell(row=r, column=25).value; z = ws.cell(row=r, column=26).value
            aa = ws.cell(row=r, column=27).value
            rounds.append({'label': lbl.strip(),
                           'points': int(y) if is_num(y) else 0,
                           'bonus': int(z) if is_num(z) else 0,
                           'total': int(aa) if is_num(aa) else 0,
                           'round': ROUND_LABELS.get(lbl.strip())})
        bonus = []
        for r in range(4, 11):
            cat = ws.cell(row=r, column=24).value; ans = ws.cell(row=r, column=25).value
            bp = ws.cell(row=r, column=29).value
            if isinstance(cat, str) and cat.strip():
                bonus.append({'category': cat.strip(),
                              'answer': ans.strip() if isinstance(ans, str) else None,
                              'points': int(bp) if is_num(bp) else 0})
        # "rätt lag vidare" — dela ut direkt baserat på vilka lag som nått varje
        # runda (1/2/4/6/8/8 p per rätt lag), oberoende av om matchen spelats.
        player_by_round = {}
        for kno, e in ko.items():
            rnd = ko_round.get(int(kno))
            if not rnd: continue
            s = player_by_round.setdefault(rnd, set())
            if e.get('predHome') in KNOWN_TEAMS: s.add(e['predHome'])
            if e.get('predAway') in KNOWN_TEAMS: s.add(e['predAway'])
        ratt_lag = {rnd: pts * len(facit_by_round.get(rnd, set()) & player_by_round.get(rnd, set()))
                    for rnd, pts in RL_PTS.items()}
        for rd in rounds:
            if rd['round'] in RL_PTS:
                rd['bonus'] = ratt_lag.get(rd['round'], 0)
                rd['total'] = rd['points'] + rd['bonus']

        # bonusfrågor: "flest mål/insläppta" kan delas av flera lag (lika räknas rätt)
        for b in bonus:
            cat = b['category'].lower()
            ans = (b['answer'] or '').strip().lower()
            if 'gör flest mål' in cat:
                b['points'] = 10 if ans and ans in scored_leaders else 0
            elif 'släpper in flest' in cat:
                b['points'] = 10 if ans and ans in conceded_leaders else 0

        by_label = {r['label']: r for r in rounds}
        gpts = by_label.get('Gruppspel', {}).get('total', 0)
        bpts = sum(b['points'] for b in bonus)
        kpts = sum(rd['total'] for rd in rounds if rd['round'] in RL_PTS)
        total = gpts + kpts + bpts
        players.append({'name': name, 'total': total, 'groupPoints': gpts,
                        'bonusPoints': bpts, 'knockoutPoints': kpts,
                        'predictions': preds, 'knockoutPreds': ko,
                        'rounds': rounds, 'bonus': bonus})
    players.sort(key=lambda p: -p['total'])

    # grupptabeller (färdigräknade i arket, rätt inbördes ordning)
    standings = {}
    for r in range(1, res.max_row + 1):
        if res.cell(row=r, column=15).value == 'Lag' and res.cell(row=r, column=19).value == 'GM':
            grp = None
            for rr in range(r, 0, -1):
                bv = res.cell(row=rr, column=2).value
                if isinstance(bv, str) and bv.startswith('Grupp '):
                    grp = bv.replace('Grupp ', '').strip(); break
            rows = []
            for tr in range(r + 1, r + 5):
                lag = res.cell(row=tr, column=15).value
                if not isinstance(lag, str) or not lag.strip():
                    continue
                def gi(c, _tr=tr):
                    v = res.cell(row=_tr, column=c).value
                    return int(v) if is_num(v) else 0
                rows.append({'team': lag.strip(), 'w': gi(16), 'd': gi(17), 'l': gi(18),
                             'gf': gi(19), 'ga': gi(20), 'gd': gi(21), 'pts': gi(22)})
            if grp and rows:
                standings[grp] = rows

    return {'updated': datetime.datetime.now(datetime.timezone.utc).isoformat(),
            'tournament': 'VM 2026',
            'matches': [matches[k] for k in sorted(matches)],
            'knockout': sorted(knockout, key=lambda m: m['no']),
            'standings': standings,
            'players': players}

def autofill_bonus(work_path, recalced_path):
    """Fyll i facit (Resultat & tabell Y5–Y6) för de bonusfrågor som är
    entydiga: världsmästare (vinnare final) och vinnare bronsmatch. Skriver bara
    i tomma celler (manuellt facit vinner).
    OBS: "flest mål/insläppta i gruppspelet" fylls INTE i automatiskt, eftersom
    flera lag ofta hamnar lika — där måste tipsgeneralen själv avgöra (Y7/Y8)."""
    rc = load_workbook(recalced_path, data_only=True)
    res = rc['Resultat & tabell']
    winners = {}  # match_no -> winning team name
    for r, no, home, away in match_rows(res):
        if no <= 72:
            continue
        hg = res.cell(row=r, column=7).value; ag = res.cell(row=r, column=9).value
        if is_num(hg) and is_num(ag) and hg != ag:
            winners[no] = home if hg > ag else away

    facit = {}
    if 104 in winners: facit[5] = winners[104]   # Y5 världsmästare (vinnare final)
    if 103 in winners: facit[6] = winners[103]   # Y6 vinnare bronsmatch

    wb = load_workbook(work_path, keep_vba=True); ws = wb['Resultat & tabell']
    written = 0
    for row, val in facit.items():
        if not val: continue
        cur = ws.cell(row=row, column=25).value
        if cur is None or (isinstance(cur, str) and not cur.strip()):
            ws.cell(row=row, column=25).value = val; written += 1
    if written:
        wb.save(work_path)
    return written


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--mock', help='lokal JSON med API-svar (för test)')
    ap.add_argument('--master', default=MASTER)
    ap.add_argument('--out', default=OUT_JSON)
    args = ap.parse_args()

    token = os.environ.get('FOOTBALL_DATA_TOKEN', '')
    if not token and not args.mock:
        print('FOOTBALL_DATA_TOKEN saknas (och ingen --mock). Avbryter.', file=sys.stderr)
        sys.exit(1)

    manual = {}
    if os.path.exists('manual_results.json'):
        manual = json.load(open('manual_results.json', encoding='utf-8'))

    matches = fetch_matches(token, args.mock)
    results = build_results(matches)
    print(f'{len(results)} spelade matcher matchade från källan.')

    # arbetskopia av master
    import shutil, openpyxl
    work = 'work.xlsm'; shutil.copy(args.master, work)

    # Pass 1: skriv gruppspel (slutspelslagen är ännu formler), räkna om
    apply_results(work, results, manual)
    recalced = recalc(work)

    # Iterera: frys framräknade slutspelslag -> skriv den rundans resultat -> räkna om.
    # Varje varv löser nästa runda (sextondel -> åttondel -> ... -> final).
    prev = -1
    for _ in range(7):
        rc = openpyxl.load_workbook(recalced, data_only=True)
        src_ws = rc['Resultat & tabell']
        wb = load_workbook(work, keep_vba=True); ws = wb['Resultat & tabell']
        for r in range(1, src_ws.max_row + 1):
            for c in (4, 6):  # hemmalag/bortalag (formler i slutspelet)
                v = src_ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip() and 'VALUE' not in v.upper():
                    cur = ws.cell(row=r, column=c).value
                    if not isinstance(cur, str) or cur.startswith('='):
                        ws.cell(row=r, column=c).value = v  # frys lagnamn för matchning
        wb.save(work)
        n = apply_results(work, results, manual)
        recalced = recalc(work)
        if n == prev:
            break
        prev = n
    print(f'Skrev {prev} matcher totalt (gruppspel + slutspel).')

    # fyll i de bonusfacit som går att räkna ut automatiskt
    if autofill_bonus(work, recalced):
        recalced = recalc(work)
        print('Fyllde i automatiskt bonusfacit (världsmästare, bronsmatchvinnare).')

    data = extract(recalced)
    json.dump(data, open(args.out, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    played = sum(1 for m in data['matches'] if m['played'])
    ko_played = sum(1 for m in data.get('knockout', []) if m['played'])
    print(f'Klart. Gruppspel {played}/{len(data["matches"])}, slutspel {ko_played}/'
          f'{len(data.get("knockout", []))}. Ledare: '
          f'{data["players"][0]["name"]} ({data["players"][0]["total"]} p).')

if __name__ == '__main__':
    main()
